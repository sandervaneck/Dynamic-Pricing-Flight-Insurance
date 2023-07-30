from openpyxl import Workbook

weather_paths = [
        "/Users/sander/Downloads/New York City 2022-01-01 to 2022-12-31.csv",
        "/Users/sander/Downloads/Orlando 2022-01-01 to 2022-12-31.csv",
        "/Users/sander/Downloads/Denver 2022-01-01 to 2022-12-31.csv",
        "/Users/sander/Downloads/Los Angeles 2022-01-01 to 2022-12-31.csv",
        "/Users/sander/Downloads/Atlanta 2022-01-01 to 2022-12-31.csv"
       ]
flight_data_paths = [
        "/Users/sander/Library/CloudStorage/GoogleDrive-mychefsbase@gmail.com/My Drive/Mack Backup/Thesis/FlightData/012022Input.csv",
        "/Users/sander/Library/CloudStorage/GoogleDrive-mychefsbase@gmail.com/My Drive/Mack Backup/Thesis/FlightData/022022Input.csv",
        "/Users/sander/Library/CloudStorage/GoogleDrive-mychefsbase@gmail.com/My Drive/Mack Backup/Thesis/FlightData/032022Input.csv",
        "/Users/sander/Library/CloudStorage/GoogleDrive-mychefsbase@gmail.com/My Drive/Mack Backup/Thesis/FlightData/042022Input.csv",
        "/Users/sander/Library/CloudStorage/GoogleDrive-mychefsbase@gmail.com/My Drive/Mack Backup/Thesis/FlightData/052022Input.csv",
        "/Users/sander/Library/CloudStorage/GoogleDrive-mychefsbase@gmail.com/My Drive/Mack Backup/Thesis/FlightData/062022Input.csv",
        "/Users/sander/Library/CloudStorage/GoogleDrive-mychefsbase@gmail.com/My Drive/Mack Backup/Thesis/FlightData/072022Input.csv",
        "/Users/sander/Library/CloudStorage/GoogleDrive-mychefsbase@gmail.com/My Drive/Mack Backup/Thesis/FlightData/082022Input.csv",
        "/Users/sander/Library/CloudStorage/GoogleDrive-mychefsbase@gmail.com/My Drive/Mack Backup/Thesis/FlightData/092022Input.csv",
        "/Users/sander/Library/CloudStorage/GoogleDrive-mychefsbase@gmail.com/My Drive/Mack Backup/Thesis/FlightData/102022Input.csv",
        "/Users/sander/Library/CloudStorage/GoogleDrive-mychefsbase@gmail.com/My Drive/Mack Backup/Thesis/FlightData/112022Input.csv",
        "/Users/sander/Library/CloudStorage/GoogleDrive-mychefsbase@gmail.com/My Drive/Mack Backup/Thesis/FlightData/122022Input.csv"
    ]

headers = ['date', 'distance', 'temp', 'dew', 'humidity', 'precip', 'windgust', 'windspeed', 'visibility','delay', 'refund', 'carrier', 'origin', 'state', 'time','weekday', 'cloudcover', 'sealevelpressure']
cat_var = ['distance', 'carrier', 'origin', 'state', 'time', 'weekday']
num_var = ['scaled_origin_score', 'scaled_weekday_score', 'scaled_carrier_score', 'scaled_time_score', 'distance', 'temp', 'dew', 'humidity', 'precip','windgust', 'windspeed', 'visibility', 'cloudcover', 'sealevelpressure']
df_var = ['refund', 'scaled_origin_score', 'scaled_time_score', 'scaled_weekday_score','cloudcover','sealevelpressure','scaled_carrier_score', 'distance', 'temp', 'dew', 'humidity', 'precip','windgust', 'windspeed', 'visibility']
factors = ['scaled_origin_score', 'scaled_time_score',  'scaled_weekday_score','sealevelpressure', 'cloudcover', 'scaled_carrier_score', 'distance', 'windspeed', 'visibility', 'precip']
rfc_factors = ['scaled_origin_score','scaled_time_score', 'scaled_weekday_score','cloudcover','sealevelpressure', 'scaled_carrier_score', 'distance', 'windspeed', 'visibility', 'precip', 'temp', 'dew', 'humidity']

# , 'temp', 'dew', 'humidity', 'precip','windgust', 'windspeed', 'visibility']

workbookDf = Workbook()
dataFile = "Data.xlsx"
workbookTrainDf = Workbook()
randomForestFile = "RandomForest.xlsx"
workbookResultsDF = Workbook()
resultFile = "Result.xlsx"