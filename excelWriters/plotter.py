import matplotlib.pyplot as plt
import sns as sns
from sklearn.metrics import r2_score, explained_variance_score, precision_score, mean_absolute_error, classification_report, RocCurveDisplay, confusion_matrix,  recall_score, accuracy_score, roc_auc_score, precision_recall_curve, f1_score, mean_squared_error
import pandas as pd
import numpy as np
import openpyxl as openpyxl
from openpyxl.utils.dataframe import dataframe_to_rows
import io
def plot_balancing(Y, workbook, row, sheet, file):
    plt.figure()
    target_count = Y.value_counts()
    target_count.plot(kind='bar', title='Refund');
    plt.title(f"{len(Y.loc[Y == 1]) / len(Y)*100}% of flights Refunded")
    temp_file = f'{row}.png'
    plt.savefig(temp_file)
    plt.close()
    img = openpyxl.drawing.image.Image(temp_file)
    # image_buffer = io.BytesIO()
    # plt.savefig(image_buffer, format='png')
    plt.close()
    # image_buffer.seek(0)
    # img = openpyxl.drawing.image.Image(image_buffer)
    sheet.add_image(img, f'A{row}')
    workbook.save(file)

def plot_roc(y, hat_prob_y):
    RocCurveDisplay.from_predictions(
        y,
        hat_prob_y,
        name="ROC curve",
        color="darkorange",
    )
    plt.plot([0, 1], [0, 1], "k--", label="chance level (AUC = 0.5)")
    plt.axis("square")
    plt.xlabel("False Positive Rate")
    plt.ylabel("True Positive Rate")
    plt.title("ROC curve")
    plt.legend()
    plt.show()

def plot_confusion_matrix(hat_y, y, target_names):
    matrix = confusion_matrix(y, hat_y)
    sns.heatmap(matrix.T, square=True, annot=True, fmt="d", cbar=False,
    xticklabels=target_names, yticklabels=target_names)
    plt.xlabel("true label")
    plt.ylabel("predicted label")

def estimate_cost(hat_y, y, cost_FP, cost_FN):
    return  np.sum(np.multiply(hat_y, (1 - y)) * cost_FP) + np.sum(np.multiply((1 - hat_y), y) * cost_FN)

def plot_to_evaluate(y, y_hat, title, workbook, file):
    sheet = workbook.create_sheet(f"{title}")

    # Calculate evaluation metrics
    costs = estimate_cost(y_hat, y, 1, 1)
    mse = mean_squared_error(y, y_hat)
    rmse = mse ** 0.5
    accuracy = accuracy_score(y, y_hat)
    classification_rep = classification_report(y, y_hat, output_dict=True)

    # Create the DataFrame for the first table
    table_data = {'Metrics': ['Costs', 'MSE', 'RMSE', 'Accuracy Score'],
                  'Values': [costs, mse, rmse, accuracy]}
    df_table = pd.DataFrame(table_data)

    # Write the first table to the sheet
    for row in dataframe_to_rows(df_table, index=False, header=True):
        sheet.append(row)

    # Create the DataFrame for the classification report
    df_classification = pd.DataFrame(classification_rep).transpose()

    # Write the second table (classification report) to the sheet
    sheet.append([])  # Add an empty row for spacing
    sheet.append(['Classification Report'])
    for row in dataframe_to_rows(df_classification, index=True, header=True):
        sheet.append(row)

    # Save the Excel workbook
    workbook.save(file)

def print_results(y, y_hat, title):
  print(f"{title} MAE:{mean_absolute_error(y, y_hat)}")
  print(f"{title} MSE:{mean_squared_error(y, y_hat)}")
  print(f"{title} RMSE:{mean_squared_error(y, y_hat)**0.5}")
  print(f"{title} Explained Variance Score:{explained_variance_score(y, y_hat)}")
  print(f"{title} R2: {r2_score(y, y_hat)}")

def plot_precision_recall_curve(probs, test_Y, label, workbook, file):
    sheet = workbook.create_sheet(f"Precision_recall_curve")
    plt.figure()
    precision, recall, _ = precision_recall_curve(test_Y, probs)
    no_skill = len(test_Y[test_Y == 1]) / len(test_Y)
    plt.plot([0, 1], [no_skill, no_skill], linestyle='--', label='No Skill')
    plt.plot(recall, precision, marker='.', label=label)
    plt.xlabel('Recall')
    plt.ylabel('Precision')
    plt.legend()
    temp_file = f'{label}.png'
    plt.title(f"Precision Recall Curve")
    plt.savefig(temp_file)
    plt.close()
    img = openpyxl.drawing.image.Image(temp_file)
    sheet.add_image(img, f'A1')
    workbook.save(file)

def plot_optimized_characteristics(tuned_delayed_model, X_val, workbook, file):
    sheet = workbook.create_sheet('Optimized Characteristics')
    best_params = tuned_delayed_model.best_params_
    sheet.append(['Best Parameters'])
    for key, value in best_params.items():
        sheet.append([key, value])

    # Get the feature importances
    importances = tuned_delayed_model.best_estimator_.feature_importances_

    # Create a DataFrame for the feature importances
    feature_importances_df = pd.DataFrame({'Feature': X_val.columns, 'Importance': importances})

    # Write the feature importances to the sheet
    sheet.append([])  # Add an empty row for spacing
    sheet.append(['Feature', 'Importance'])
    for row in dataframe_to_rows(feature_importances_df, index=False, header=True):
        sheet.append(row)

    # Plot the feature importances
    plt.figure()
    temp_file = f'../optimized.png'
    plt.title(f"Feature Importances")
    plt.bar(X_val.columns, importances)  # Use the feature names on the x-axis
    plt.xticks(rotation='vertical')  # Rotate the x-axis labels vertically for better readability
    plt.savefig(temp_file)
    plt.close()

    # Add the image to the sheet
    img = openpyxl.drawing.image.Image(temp_file)
    sheet.add_image(img, f'A1')

    # Save the Excel workbook
    workbook.save(file)


def print_df_overview(df, workbook, file, cat_var):
    for index, row in df.iterrows():
        if 'time' not in row:
            date_value = row.get('date')  # Assuming 'date' is the column name for the date
            state_name_value = row.get('state')  # Assuming 'state_name' is the column name for the state name
            print(f"Time not found at index {index}. Date: {date_value}, State Name: {state_name_value}")
    plot_timeline_refunds(df, workbook.create_sheet('Timeline'))
    sheet = workbook.create_sheet('Variables')
    headers = cat_var
    values = []
    for var in cat_var:
        values.append(len(df[var].unique()))
    for row in zip(headers, values):
        sheet.append(row)
    sheet.append(['#Flights', len(df)])
    # Save the Excel workbook
    workbook.save(file)

def plot_timeline_refunds(df, wb, sheet, file):
    plt.figure()
    # df['date'] = df['date'].dropna()
    fig, ax1 = plt.subplots(figsize=(30, 5))
    df.set_index('date', inplace=True)
    ax1.plot(df.groupby('date')['refund'].mean(), data=df, color='g')
    ax1.set_xlabel('Days of the year')
    ax1.set_ylabel('Refund %', color='g')
    # ax1.set_xlim(min(df['date']), max(df['date']))
    temp_file = f'../distr timeline.png'
    plt.savefig(temp_file)
    plt.close()
    img = openpyxl.drawing.image.Image(temp_file)
    sheet.add_image(img, f'A1')
    wb.save(file)

def plot_timeline_test(y_hat, y, wb, sheet, file):
    plt.figure()
    # df['date'] = df['date'].dropna()
    fig, ax1 = plt.subplots(figsize=(30, 5))
    df.set_index('date', inplace=True)
    ax1.plot(df.groupby('date')['refund'].mean(), data=df, color='g')
    ax1.set_xlabel('Days of the year')
    ax1.set_ylabel('Refund %', color='g')
    # ax1.set_xlim(min(df['date']), max(df['date']))
    temp_file = f'../distr timeline.png'
    plt.savefig(temp_file)
    plt.close()
    img = openpyxl.drawing.image.Image(temp_file)
    sheet.add_image(img, f'A1')
    wb.save(file)
