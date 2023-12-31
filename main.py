import warnings

import numpy as np
import openpyxl
import optuna
import pandas as pd
import seaborn as sns
import statsmodels.api as sm
import statsmodels.formula.api as smf
from imblearn.metrics import specificity_score
from matplotlib import pyplot as plt
from openpyxl.workbook import Workbook
from sklearn.ensemble import RandomForestClassifier, GradientBoostingClassifier, BaggingClassifier
from sklearn.inspection import permutation_importance
from sklearn.linear_model import LogisticRegression
from sklearn.metrics import make_scorer, f1_score, precision_score, accuracy_score, recall_score
from sklearn.model_selection import train_test_split, cross_val_score
from sklearn.svm import SVC
from sklearn.tree import plot_tree

from constants.paths import weather_paths
from dataHandling.addCatScores import add_cat_scores
from dataHandling.createdf import parse_data
from evaluations.crossvalidation import crossvalscore, crossvalscore_glm
from evaluations.feature_importances import feature_importances
from resample import resample_combination
from tuning.objective import total_score

# Press the green button in the gutter to run the script.

if __name__ == '__main__':

    # Ignore all warnings
    warnings.filterwarnings("ignore")
    wb = Workbook()
    wb1 = Workbook()
    wb2 = Workbook()
    file = "File.xlsx"
    file2 = "File2.xlsx"
    file3 = "File3.xlsx"

    weathers = []
    summarized_data = []

    df = parse_data(summarized_data, weather_paths, weathers)
    df.dropna()

    df['distance'] = df['distance'].astype(float)
    df = add_cat_scores(df, file2, wb2)

    refunded_df = df[df['refund'] == 1]
    non_refunded_df = df[df['refund'] == 0]
    plt.figure(figsize=(10, 8))
    target_count = refunded_df['weekday'].value_counts()

    target_count.plot(figsize=(5, 5), kind="bar", title=f"Frequency distribution variable weekday")
    plt.show()

    plt.figure(figsize=(10, 8))
    target_count2 = non_refunded_df['weekday'].value_counts()

    target_count2.plot(figsize=(5, 5), kind="bar", title=f"Frequency distribution variable weekday")
    plt.show()
    #




    df['wind_precip'] = df['windspeed'] * df['precip']
    df['visibility_squared'] = df['visibility'] * df['visibility']

    scores_vars = ['scaled_carrier_score', 'scaled_origin_score', 'scaled_time_score', 'scaled_weekday_score']
    time_var = ['date']
    numerical_vars = ['distance', 'windspeed', 'visibility', 'temp', 'sealevelpressure', 'precip', 'cloudcover']

    corr_matrix = df[scores_vars + numerical_vars].corr()

    # Set up the matplotlib figure
    plt.figure(figsize=(10, 8))

    # Create a heatmap of the correlation matrix
    # You can customize the colormap (cmap) as needed
    sns.heatmap(corr_matrix, annot=True, cmap='coolwarm', vmin=-1, vmax=1)

    # Add plot title
    plt.title('Correlation Matrix')

    # Show the plot
    plt.show()

    nonlinear_vars = ['wind_precip', 'visibility_squared']
    sheet4 = wb1.create_sheet("timeline")
    # plot_timeline_refunds(df, wb1, sheet4, file3 )
    xs = scores_vars + numerical_vars
    xs_nonlinear = scores_vars + nonlinear_vars + ['temp', 'sealevelpressure', 'cloudcover', 'distance']
    ys = ['refund']

    seed = 42
    train_df, aux_df = train_test_split(df, train_size=.5, random_state=seed)
    validation_df, test_df = train_test_split(aux_df, train_size=.5, random_state=seed)

    train_X = train_df.loc[:, xs]
    train_X_nonlinear = train_df.loc[:, xs_nonlinear]
    train_Y = train_df.loc[:, ys]

    print(test_df["date"].describe())

    print(train_df["refund"].value_counts())
    print(train_df["time"].value_counts())

    train_X_resampled, train_Y_resampled_ = resample_combination(train_X, train_Y, '', '')

    val_X = validation_df.loc[:, xs]
    val_Y = validation_df.loc[:, ys]
    test_X = test_df.loc[:, xs]
    test_X_non_linear = test_df.loc[:, xs_nonlinear]

    test_Y = test_df.loc[:, ys]

    formula = 'refund ~ distance + windspeed + visibility + temp + cloudcover + precip + sealevelpressure + scaled_origin_score + scaled_carrier_score + scaled_origin_score + scaled_time_score + scaled_weekday_score'
    formula_nonlinear = 'refund ~ wind_precip + visibility_squared + temp + cloudcover + distance + sealevelpressure + scaled_origin_score + scaled_carrier_score + scaled_origin_score + scaled_time_score + scaled_weekday_score'

    logistic_regression = LogisticRegression(random_state=seed)
    rand_clf = RandomForestClassifier(random_state=5)
    svm_clf = SVC(random_state=seed)
    gbm = GradientBoostingClassifier(random_state=seed)
    glm = smf.glm(formula=formula, data=train_df, family=sm.families.Binomial())
    glm_nonlinear = smf.glm(formula=formula_nonlinear, data=train_df, family=sm.families.Binomial())
    # #
    feature_importances(svm_clf, train_X, train_Y, xs, 'svm')
    feature_importances(svm_clf, train_X_resampled, train_Y_resampled_, xs, 'reb_svm')
    feature_importances(rand_clf, train_X, train_Y, xs, 'rfc')
    feature_importances(rand_clf, train_X_resampled, train_Y_resampled_, xs, 'reb_rfc')
    print('rfc done')
    feature_importances(logistic_regression, train_X, train_Y, xs, 'lr')
    feature_importances(logistic_regression, train_X_resampled, train_Y_resampled_, xs, 'reb_lr')
    print('lr done')
    feature_importances(gbm, train_X, train_Y, xs, 'gbm')
    feature_importances(gbm, train_X_resampled, train_Y_resampled_, xs, 'reb_gbm')
    print('gbm done')

    crossvalscore(logistic_regression, train_X, train_Y, 'logistic_regression')
    crossvalscore(rand_clf, train_X, train_Y, 'rand_clf')
    crossvalscore(svm_clf, train_X, train_Y, 'svm')
    crossvalscore(gbm, train_X, train_Y, 'gbm')
    crossvalscore(logistic_regression, train_X_resampled, train_Y_resampled_, 'logistic_regression_res')
    crossvalscore(rand_clf, train_X_resampled, train_Y_resampled_, 'rand_clf_res')
    crossvalscore(gbm, train_X_resampled, train_Y_resampled_, 'gbm_res')
    #
    glm_result = glm.fit()
    glm_nl_result = glm_nonlinear.fit()
    print(glm_result.summary())
    #
    glm_result_nonlinear = glm_nonlinear.fit()
    print(glm_result_nonlinear.summary())

    crossvalscore_glm(glm_result, sm.add_constant(train_X), train_df['refund'], 'glm ')
    crossvalscore_glm(glm_nl_result, sm.add_constant(train_X_nonlinear), train_df['refund'], 'glm_nonlinear ')

    study = optuna.create_study(direction="maximize")
    def objective(trial):
        classifier = trial.suggest_categorical('classifier', ['RandomForest', 'SVC', 'LogisticRegression', 'GBM'])

        if classifier == 'RandomForest':
            n_estimators = trial.suggest_int('n_estimators', 10, 300, 10)
            max_depth = int(trial.suggest_int('max_depth', 10, 100, log=True))
            min_samples_split = trial.suggest_int('min_samples_split', 2, 20)
            min_samples_leaf = trial.suggest_int('min_samples_leaf', 1, 20)

            clf = RandomForestClassifier(
                n_estimators=n_estimators,
                max_depth=max_depth,
                min_samples_split=min_samples_split,
                min_samples_leaf=min_samples_leaf
            )
        elif classifier == 'SVC':
            c = trial.suggest_float('svc_c', 0.1, 1, log=True)
            coef0 = trial.suggest_float('svc_coef0', -1.0, 1.0)
            probability = trial.suggest_categorical('svc_prob', [True, False])
            class_weight = trial.suggest_categorical('svc_class_weight', [None, 'balanced'])
            n_estimators = 100
            clf = BaggingClassifier(SVC(C=c,coef0=coef0,probability=probability, class_weight=class_weight), max_samples=1.0 / n_estimators, n_estimators=n_estimators)
        elif classifier == 'GradientBoosting':
            learning_rate = trial.suggest_float('gb_learning_rate', 0.01, 1.0, log=True)
            n_estimators = trial.suggest_int('gb_n_estimators', 50, 500, 50)
            max_depth = trial.suggest_int('gb_max_depth', 3, 10)
            subsample = trial.suggest_float('gb_subsample', 0.1, 1.0)
            clf = GradientBoostingClassifier(
                learning_rate=learning_rate,
                n_estimators=n_estimators,
                max_depth=max_depth,
                subsample=subsample
            )
        else:
            C = trial.suggest_float('logreg_C', 1e-5, 1e5, log=True)
            solver = trial.suggest_categorical('logreg_solver', ['newton-cg',  'liblinear', 'saga'])
            if solver in ['newton-cg', 'lbfgs']:
                penalty = trial.suggest_categorical('logreg_penalty', ['l2'])
            elif solver == 'liblinear':
                penalty = trial.suggest_categorical('logreg_penalty', ['l2'])
            else:
                penalty = 'l1'

            max_iter = trial.suggest_int('logreg_max_iter', 100, 3000, 100)

            clf = LogisticRegression(C=C, solver=solver, penalty=penalty, max_iter=max_iter)

        f1_scorer = make_scorer(f1_score, average='weighted')
        recall_scorer = make_scorer(recall_score, average='weighted')
        precision_scorer = make_scorer(precision_score, average='weighted', zero_division=0)
        accuracy_scorer = make_scorer(accuracy_score)
        specificity = make_scorer(specificity_score)

        if classifier == "RandomForest":
            y = np.ravel(val_Y)
        else:
            y = val_Y
        recall_score_mean = cross_val_score(clf, val_X, np.ravel(y), n_jobs=-1, cv=5, scoring=recall_scorer).mean()
        f1_score_mean = cross_val_score(clf, val_X, np.ravel(y), n_jobs=-1, cv=5, scoring=f1_scorer).mean()
        precision_mean = cross_val_score(clf, val_X, np.ravel(y), n_jobs=-1, cv=5, scoring=precision_scorer).mean()
        accuracy_mean = cross_val_score(clf, val_X, np.ravel(y), n_jobs=-1, cv=5, scoring=accuracy_scorer).mean()
        specificty_mean = cross_val_score(clf, val_X, np.ravel(y), n_jobs=-1, cv=5, scoring=specificity).mean()
        rmse_mean = np.sqrt(-cross_val_score(clf, val_X, np.ravel(y), cv=5, scoring='neg_mean_squared_error', n_jobs=-1)).mean()
        auc_mean = cross_val_score(clf, val_X, np.ravel(y), n_jobs=-1, cv=5, scoring='roc_auc').mean()
        return total_score(auc_mean, accuracy_mean, precision_mean, recall_score_mean, f1_score_mean, specificty_mean, rmse_mean)

    study.optimize(objective, n_trials=100)
    trial = study.best_trial
    best_params = trial.params
    print("Best Hyperparameters: {}".format(best_params))

    # Use the best_params to create the best model
    if best_params['classifier'] == 'RandomForest':
        best_model = RandomForestClassifier(
            n_estimators=best_params['n_estimators'],
            max_depth=best_params['max_depth'],
            min_samples_split=best_params['min_samples_split'],
            min_samples_leaf=best_params['min_samples_leaf'],
        )
    elif best_params['classifier'] == 'SVC':
        best_model = SVC(
            C=best_params['svc_c'],
            tol=best_params['svc_tol'],
            degree=best_params['svc_degree'],
            coef0=best_params['svc_coef0'],
            gamma='auto'
        )
    elif best_params['classifier'] == 'GradientBoosting':
        learning_rate = trial.suggest_float('gb_learning_rate', 0.01, 1.0, log=True)
        n_estimators = trial.suggest_int('gb_n_estimators', 50, 500, 50)
        max_depth = trial.suggest_int('gb_max_depth', 3, 10)
        subsample = trial.suggest_float('gb_subsample', 0.1, 1.0)
        best_model = GradientBoostingClassifier(
            learning_rate=best_params['gb_learning_rate'],
            n_estimators=best_params['gb_n_estimators'],
            max_depth=best_params['gb_max_depth'],
            subsample=best_params['gb_subsample'],
        )
    else:
        best_model = LogisticRegression(
            C=best_params['logreg_C'],
            solver=best_params['logreg_solver'],
            max_iter=best_params['logreg_max_iter'],
            # penalty=best_params['logreg_penalty']
        )

    if best_params['classifier'] == 'RandomForest':
        y = np.ravel(test_Y)
    else:
        y = test_Y

    predicted_outcomes = best_model.predict(test_X)

    # Plot observed outcomes and predicted outcomes
    plt.figure(figsize=(10, 6))
    plt.plot(test_Y['date'], test_Y['refund'], label='Observed Outcomes')
    plt.plot(test_Y['date'], predicted_outcomes, label='Predicted Outcomes', linestyle='dashed')
    plt.xlabel('Date')
    plt.ylabel('Refund')
    plt.title('Observed vs. Predicted Outcomes')
    plt.legend()
    temp_file = f'plot_results.png'
    plt.savefig(temp_file)
    plt.close()

    scaled_predicted_outcomes = 134.68 + 55.62 * predicted_outcomes

    # Plot observed outcomes and scaled predicted outcomes
    plt.figure(figsize=(10, 6))
    plt.plot(test_Y['date'], test_Y['refund'], label='Observed Outcomes')
    plt.plot(test_Y['date'], scaled_predicted_outcomes, label='Scaled Predicted Outcomes', linestyle='dashed')
    plt.xlabel('Date')
    plt.ylabel('Scaled Refund')
    plt.title('Observed vs. Scaled Predicted Outcomes')
    plt.legend()
    temp_file = f'plot_scaled_results.png'
    plt.savefig(temp_file)
    plt.close()
    best_model.fit(val_X, val_Y)
    print(crossvalscore(best_model, test_X, y, 'ml_test'))

    perm_importance = permutation_importance(best_model, test_X, y)

    features = np.array(xs)

    sorted_indices = perm_importance.importances_mean.argsort()[::-1]
    sorted_features = features[sorted_indices]
    sorted_importances = perm_importance.importances_mean[sorted_indices]

    # Print the importances with feature names
    for feature, importance in zip(sorted_features, sorted_importances):
        print(f"{feature}: {importance}")

    # Plot the feature importances in a bar chart
    plt.figure(figsize=(10, 6))
    plt.bar(sorted_features, sorted_importances, color='skyblue')
    plt.xlabel('Feature')
    plt.ylabel('Permutation Importance')
    plt.title('Feature Importances for Tuned Model')
    plt.xticks(rotation=45, ha='right')
    plt.tight_layout()
    temp_file = f'feature Importances.png'
    plt.savefig(temp_file)
    img = openpyxl.drawing.image.Image(temp_file)
    plt.close()

    atlantaScore = df.loc[df['origin'] == 'ATL', 'scaled_origin_score'].values[0]
    orlandoScore = df.loc[df['origin'] == 'MCO', 'scaled_origin_score'].values[0]
    miamiScore = df.loc[df['origin'] == 'MIA', 'scaled_origin_score'].values[0]

    morningScore = df.loc[df['time'] == 1, 'scaled_time_score']
    if not morningScore.empty:
        # If there are matching rows, get the eveningScore
        morningScore = morningScore.values[0]
    else:
        # If there are no matching rows, assign a default value or handle as needed
        evening_score_value = 0
    afternoonScore = df.loc[df['time'] == 2, 'scaled_time_score']
    if not afternoonScore.empty:
        # If there are matching rows, get the eveningScore
        afternoonScore = afternoonScore.values[0]
    else:
        # If there are no matching rows, assign a default value or handle as needed
        evening_score = 0
    evening_score = df.loc[df['time'] == 3, 'scaled_time_score']
    if not evening_score.empty:
        # If there are matching rows, get the eveningScore
        evening_score = evening_score.values[0]
    else:
        # If there are no matching rows, assign a default value or handle as needed
        evening_score = 0
    # eveningScore = df.loc[df['time'] == 2, 'scaled_time_score'].values[0]
    nightScore = df.loc[df['time'] == 4, 'scaled_time_score']
    if not nightScore.empty:
        # If there are matching rows, get the eveningScore
        nightScore = nightScore.values[0]
    else:
        # If there are no matching rows, assign a default value or handle as needed
        nightScore = 0

    carrierYVScore = df.loc[df['carrier'] == 'YV', 'scaled_carrier_score'].values[0]
    carrierG4Score = df.loc[df['carrier'] == 'F9', 'scaled_carrier_score'].values[0]
    carrierQXScore = df.loc[df['carrier'] == 'QX', 'scaled_carrier_score'].values[0]

    mondayScore = df.loc[df['weekday'] == 'Monday', 'scaled_weekday_score'].values[0]
    tuesdayScore = df.loc[df['weekday'] == 'Tuesday', 'scaled_weekday_score'].values[0]
    wednesdayScore = df.loc[df['weekday'] == 'Wednesday', 'scaled_weekday_score'].values[0]
    saturdayScore = df.loc[df['weekday'] == 'Saturday', 'scaled_weekday_score'].values[0]

    def quantilesFor(title):
       return df[title].quantile(0.10), df[title].quantile(0.50), df[title].quantile(0.90)

    wind_10th, windh_50th, wind_90th = quantilesFor('windspeed')
    vis_10th, vis_50th, vis_90th = quantilesFor('visibility')
    vis_10th_squared = (vis_10th*vis_10th).round(3)
    vis_50th_squared = (vis_50th*vis_50th).round(3)
    vis_90th_squared = (vis_10th*vis_90th).round(3)

    temp_10th, temp_50th, temp_90th = quantilesFor('temp')
    pres_10th, pres_50th, pres_90th = quantilesFor('sealevelpressure')
    dist_10th, dist_50th, dist_90th = quantilesFor('distance')
    precip_10th, precip_50th, precip_90th = quantilesFor('precip')
    cloud_10th, cloud_50th, cloud_90th = quantilesFor('cloudcover')
    wind_press_90th = (wind_90th * precip_90th).round(3)
    wind_press_10th = (wind_10th * precip_10th).round(3)
    wind_press_50th = (windh_50th * precip_50th).round(3)

    delay_point = [carrierYVScore,orlandoScore,afternoonScore,tuesdayScore,dist_90th,wind_90th,vis_10th,temp_10th,pres_10th, precip_90th, cloud_90th]
    delay_point_nonlinear = [carrierYVScore,orlandoScore,afternoonScore,tuesdayScore, wind_press_90th, vis_10th_squared, temp_10th, pres_10th, cloud_90th, dist_90th]
    average_point = [carrierG4Score, miamiScore, afternoonScore, tuesdayScore, dist_50th, windh_50th, vis_50th, temp_50th, pres_50th, precip_50th, cloud_50th]
    average_point_nonlinear = [carrierG4Score, miamiScore, afternoonScore, tuesdayScore, wind_press_50th, vis_50th_squared, temp_50th, pres_50th, cloud_50th, dist_50th]
    nondelay_point = [carrierQXScore, atlantaScore, nightScore, tuesdayScore, dist_10th, wind_10th, vis_90th, temp_90th, pres_90th, precip_10th, cloud_10th]
    nondelay_point_nonlinear = [carrierQXScore, atlantaScore, nightScore, tuesdayScore, wind_press_10th, vis_90th_squared, temp_90th, pres_90th, cloud_10th, dist_10th]

    print("Delay Point")
    print("Predicted RF:", best_model.predict_proba(pd.DataFrame([delay_point], columns=xs)))
    print("Predicted GLM:", glm_result.predict(pd.DataFrame([delay_point], columns=xs)))
    print("Predicted NonLinear_GLM:", glm_nl_result.predict(pd.DataFrame([delay_point_nonlinear], columns=xs_nonlinear)))

    print("Average Point")
    print("Predicted RF:", best_model.predict_proba(pd.DataFrame([average_point], columns=xs)))
    print("Predicted GLM:", glm_result.predict(pd.DataFrame([average_point], columns=xs)))
    print("Predicted NonLinear_GLM:", glm_nl_result.predict(pd.DataFrame([average_point_nonlinear], columns=xs_nonlinear)))

    print("Non delay Point")
    print("Predicted RF:", best_model.predict_proba(pd.DataFrame([nondelay_point], columns=xs)))
    print("Predicted GLM:", glm_result.predict(pd.DataFrame([nondelay_point], columns=xs)))
    print("Predicted NonLinear_GLM:", glm_nl_result.predict(pd.DataFrame([nondelay_point_nonlinear], columns=xs_nonlinear)))

    test_X_c = sm.add_constant(test_X)
    test_X_c_non_linear = sm.add_constant(test_X_non_linear)

    # Use the fitted GLM model to make predictions on the test data
    crossvalscore_glm(glm_result,test_X_c, test_df['refund'], 'glm predictions')
    crossvalscore_glm(glm_nl_result,test_X_c_non_linear, test_df['refund'], 'glm_nonlinear predictions')

    tree_index = 0

    # Get the specific decision tree from the random forest
    decision_tree = best_model.estimators_[tree_index]

    # Plot the decision tree
    plt.figure(figsize=(20, 10))
    plot_tree(decision_tree, filled=True, feature_names=xs, precision=1, max_depth=3)  # Adjust max_depth for visibility
    plt.tight_layout()
    temp_file = f'tree.png'
    plt.savefig(temp_file)
    img = openpyxl.drawing.image.Image(temp_file)
    plt.close()

    # Get the specific decision tree from the random forest
    tree_index = 0  # You can change this if you want to select a different tree
    decision_tree = best_model.estimators_[tree_index]

