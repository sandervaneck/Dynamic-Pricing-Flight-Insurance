from tabulate import tabulate
import io
import matplotlib.pyplot as plt
import openpyxl as openpyxl
from resample import plot_balancing

def create_table(df, factors):
    statistics = df.groupby('refund').describe()
    statistics_filtered = statistics[factors].transpose()
    statistics_filtered['delta'] = statistics_filtered.iloc[:, 0] - statistics_filtered.iloc[:, 1]
    table = tabulate(statistics_filtered, headers='keys', tablefmt='fancy_grid')
    return table

def plot_distributions(df, cat_var, sheet, cell):
    for i, var in enumerate(cat_var):
        plt.figure()
        target_count = df[var].value_counts()
        if var != 'distance':
            target_count.plot(figsize=(5, 5), kind="bar", title=f"Frequency distribution variable {var}")
        else:
            ax = target_count.plot(figsize=(5, 5), kind="bar", title=f"Frequency distribution variable {var}")
            ax.set_xticklabels([])
        image_buffer = io.BytesIO()
        plt.savefig(image_buffer, format='png')
        plt.close()
        image_buffer.seek(0)
        img = openpyxl.drawing.image.Image(image_buffer)
        sheet.add_image(img, f'{cell}{i * 25 + 1}')



def write_table(df, numerical_var, sheet):
    statistics = df.groupby('refund').describe()
    statistics_filtered = statistics[numerical_var].transpose()
    statistics_filtered['delta'] = statistics_filtered.iloc[:, 0] - statistics_filtered.iloc[:, 1]

    # Add variable name and descriptive type to the first column
    statistics_filtered.insert(0, 'Variable', statistics_filtered.index)
    # statistics_filtered.insert(1, 'Descriptive Type', ['mean', 'percentile'] * (len(statistics_filtered) // 2))

    # Convert DataFrame to a list of lists
    table_data = [list(statistics_filtered.columns)]
    table_data += statistics_filtered.values.tolist()

    # Write table to the sheet
    for row_index, row_data in enumerate(table_data, start=1):
        for col_index, cell_value in enumerate(row_data, start=1):
            # Convert tuple to string if the value is a tuple
            if isinstance(cell_value, tuple):
                cell_value = ', '.join(str(value) for value in cell_value)
            sheet.cell(row=row_index, column=col_index, value=cell_value)

def show_descriptives(df, cat_var, numerical_var, workbook, file):
    write_table(df, numerical_var, workbook.create_sheet(title='Descriptives'))

    refunded_df = df[df['refund'] == 1]
    non_refunded_df = df[df['refund'] == 0]

    sheet_for_categorical_distributions = workbook.create_sheet("Categorical variables")
    plot_distributions(
        non_refunded_df,
        cat_var,
        sheet_for_categorical_distributions,
        'A')
    plot_distributions(
        refunded_df,
        cat_var,
        sheet_for_categorical_distributions,
        'K')

    plot_numerical_distribution(
        workbook.create_sheet('Numerical Variables'),
        refunded_df,
        non_refunded_df,
        numerical_var
         )

    plot_balancing(df['refund'], workbook, 1, workbook.create_sheet("Balance"), file)

def plot_numerical_distribution(sheet, refunded_df, non_refunded_df, numerical_var):
    refunded_df.set_index('date', inplace=True)
    non_refunded_df.set_index('date', inplace=True)

    refunded_avg = refunded_df[numerical_var].groupby(refunded_df.index).mean()
    non_refunded_avg = non_refunded_df[numerical_var].groupby(non_refunded_df.index).mean()

    # Increase Agg rendering parameters
    plt.rcParams['agg.path.chunksize'] = 1000
    plt.rcParams['path.simplify_threshold'] = 0.01

    # Define colors for refunded and non-refunded values
    refunded_color = 'blue'
    non_refunded_color = 'red'

    # Plot line graphs for each variable
    for i, column in enumerate(numerical_var):
        if column not in ['date', 'delay']:
            plt.figure()
            plt.plot(refunded_avg.index, refunded_avg[column], color=refunded_color, label='Refunded', linestyle='-', linewidth=1)
            plt.plot(non_refunded_avg.index, non_refunded_avg[column], color=non_refunded_color, label='Non-Refunded', linestyle='-', linewidth=1)
            plt.xlabel('Date')
            plt.ylabel(f'Average {column}')
            plt.title(f'Average {column} over time')
            # Save the image to a buffer in memory instead of a file
            image_buffer = io.BytesIO()
            plt.savefig(image_buffer, format='png')
            plt.close()
            image_buffer.seek(0)
            img = openpyxl.drawing.image.Image(image_buffer)

            # Add the image to the Excel sheet
            sheet.add_image(img, f'A{i * 24 + 1}')
